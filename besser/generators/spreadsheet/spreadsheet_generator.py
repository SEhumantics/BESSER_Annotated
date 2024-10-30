import os
from besser.BUML.metamodel.structural import DomainModel, Class, BinaryAssociation
from besser.generators import GeneratorInterface
from besser.utilities import sort_by_timestamp
from openpyxl import Workbook
from openpyxl.styles import Font, NamedStyle
from openpyxl.worksheet.datavalidation import DataValidation


class SpreadSheetGenerator(GeneratorInterface):
    """
    Generates an Excel spreadsheet from a UML class diagram model.
    Each class creates a sheet, each attribute creates a column, and associations create comboboxes.
    """
    
    # Define styles and example values
    MAX_ROWS = 3  # Define the max rows for the data

    text_style = NamedStyle(name="text_style", number_format="General")
    number_style = NamedStyle(name="number_style", number_format="0.00")
    date_style = NamedStyle(name="date_style", number_format="DD/MM/YYYY")
    time_style = NamedStyle(name="time_style", number_format="HH:MM:SS")
    datetime_style = NamedStyle(name="datetime_style", number_format="DD/MM/YYYY HH:MM:SS")

    TYPES = {
        "int": [number_style, "10"],
        "str": [text_style, "text"],
        "float": [number_style, 10.5],
        "bool": [text_style, True],
        "time": [time_style, "10:15:02"],
        "date": [date_style, "31/12/2020"],
        "datetime": [datetime_style, "31/12/2020 10:15:02"]
    }

    def __init__(self, model: DomainModel, output_dir: str = None):
        super().__init__(model, output_dir)
        self.__workbook = None

    def generate(self, *args) -> None:
        """
        Main method to generate the spreadsheet. Loops through classes to create sheets and columns,
        and sets data validation for associations.
        """
        self.__workbook = Workbook()

        # Add styles to workbook to prevent re-creation issues
        for style in (self.text_style, self.number_style, self.date_style, self.time_style, self.datetime_style):
            if style.name not in self.__workbook.named_styles:
                self.__workbook.add_named_style(style)

        # Remove default sheet created by openpyxl
        default_sheet = self.__workbook["Sheet"]
        self.__workbook.remove(default_sheet)

        self._create_class_sheets()
        self._create_association_comboboxes()
        self._create_bridge_sheets()

        # Save file
        file_path = self.build_generation_path(file_name="model.xlsx")
        self.__workbook.save(file_path)
        print(f"Generated file: '{file_path}'")

    def _create_class_sheets(self) -> None:
        """
        Creates a sheet for each class, applies styles, and fills with example values.
        """
        for buml_class in sort_by_timestamp(self.model.get_classes()):
            sheet = self.__workbook.create_sheet(title=buml_class.name)

            attrs = sort_by_timestamp(buml_class.attributes)
            for column_index, attribute in enumerate(attrs, start=1):
                # Header cell with bold font
                header_cell = sheet.cell(row=1, column=column_index, value=attribute.name)
                header_cell.font = Font(bold=True)

                # Apply styles and example values to the data rows
                style, example_value = self.TYPES.get(attribute.type.name, (self.text_style, "Example"))
                for row in range(2, self.MAX_ROWS):
                    data_cell = sheet.cell(row=row, column=column_index)
                    data_cell.style = style
                    if row == 2:  # Set example value only in the first row for example data
                        if style == self.text_style:
                            example_value = f"{buml_class.name}_{attribute.name}"
                        data_cell.value = example_value

    def _create_association_comboboxes(self) -> None:
        """
        Adds comboboxes to each class sheet for one-to-one and one-to-many associations, linked to
        other sheets.
        """
        for buml_class in sort_by_timestamp(self.model.get_classes()):
            for end in sort_by_timestamp(buml_class.association_ends()):
                if end.multiplicity.max == 1:
                    sheet = self.__workbook[buml_class.name]
                    new_column_index = sheet.max_column + 1
                    header_cell = sheet.cell(row=1, column=new_column_index, value=end.name)
                    header_cell.font = Font(bold=True)

                    # Set the formula to reference the options in the associated class sheet
                    formula = f"{end.type.name}!$A$2:$A$50"
                    data_validation = DataValidation(type="list", formula1=formula, showDropDown=False)
                    data_validation.error = "Select a valid option"
                    data_validation.errorTitle = "Invalid option"

                    # Apply the data validation to each cell in the new column
                    for row in range(2, self.MAX_ROWS):
                        cell = sheet.cell(row=row, column=new_column_index)
                        data_validation.add(cell)

                    # Add data validation to the sheet
                    sheet.add_data_validation(data_validation)

                    # Automatically set the first value of the dropdown for the first row
                    first_option = self.__workbook[end.type.name].cell(row=2, column=1).value
                    sheet.cell(row=2, column=new_column_index, value=first_option)

    def _configure_association_column(self, sheet, end, column_index) -> None:
        """Configures a column in the sheet for an association end with data validation."""
        header_cell = sheet.cell(row=1, column=column_index, value=end.name)
        header_cell.font = Font(bold=True)
        formula = f"{end.type.name}!$A$2:$A$50"
        data_validation = DataValidation(type="list", formula1=formula, showDropDown=False)
        data_validation.error = "Select a valid option"
        data_validation.errorTitle = "Invalid option"

        for row in range(2, self.MAX_ROWS):
            cell = sheet.cell(row=row, column=column_index)
            data_validation.add(cell)
        
        sheet.add_data_validation(data_validation)

        # Set default value for the first row
        first_option = self.__workbook[end.type.name].cell(row=2, column=1).value
        sheet.cell(row=2, column=column_index, value=first_option)

    def _create_bridge_sheets(self) -> None:
        """
        Creates a sheet for many-to-many relationships.
        """
        for assoc in self.model.associations:
            if not isinstance(assoc, BinaryAssociation):
                continue
            if all(end.multiplicity.max == 9999 for end in assoc.ends):
                sheet = self.__workbook.create_sheet(title=assoc.name)

                # Get the sorted ends
                end1, end2 = sort_by_timestamp(assoc.ends)
                # Configure each association column
                self._configure_association_column(sheet, end1, 1)
                self._configure_association_column(sheet, end2, 2)
